"""Parser del 'Formulario apuesta' (Excel .xlsm) del Mundial 2026.

El formato tiene 12 grupos (A–L), cada uno con 6 partidos en columnas:
    C = equipo local | D = goles local | E = goles visitante | F = equipo visitante
y al final un bloque de pronóstico de los 4 primeros puestos:
    B = '1 er puesto' | E = puntos | F = equipo
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


@dataclass
class ParsedMatch:
    grupo: str
    local: str
    visitante: str
    goles_local: int
    goles_visitante: int


@dataclass
class ParsedPosition:
    posicion: int
    equipo: str
    puntos: int


@dataclass
class ParsedFormulario:
    matches: list[ParsedMatch] = field(default_factory=list)
    posiciones: list[ParsedPosition] = field(default_factory=list)

    @property
    def grupos(self) -> dict[str, list[str]]:
        """Equipos por grupo en orden de aparición."""
        out: dict[str, list[str]] = {}
        for m in self.matches:
            teams = out.setdefault(m.grupo, [])
            for t in (m.local, m.visitante):
                if t not in teams:
                    teams.append(t)
        return out


def _clean(value) -> str:
    return str(value).strip() if value is not None else ""


def parse_formulario(source: str | Path | bytes) -> ParsedFormulario:
    """Lee el formulario desde una ruta o bytes y devuelve los datos parseados."""
    if isinstance(source, (bytes, bytearray)):
        wb = openpyxl.load_workbook(BytesIO(source), data_only=True)
    else:
        wb = openpyxl.load_workbook(source, data_only=True)

    ws = wb["Formulario apuesta"] if "Formulario apuesta" in wb.sheetnames else wb.active

    result = ParsedFormulario()
    current_group: str | None = None

    for row in ws.iter_rows():
        vals = {get_column_letter(idx): cell.value for idx, cell in enumerate(row, start=1)}

        # ¿Encabezado de grupo? (p. ej. "   Grupo A" en columna C o D)
        for col in ("C", "D", "B"):
            v = vals.get(col)
            if isinstance(v, str) and "grupo" in v.lower():
                letter = re.sub(r"(?i)grupo", "", v).strip()
                if letter:
                    current_group = letter.split()[0].upper()
                break

        # ¿Fila de partido? local(C) visitante(F) con goles numéricos (D,E)
        c, f, gl, gv = vals.get("C"), vals.get("F"), vals.get("D"), vals.get("E")
        if (
            current_group
            and isinstance(c, str)
            and isinstance(f, str)
            and "sentido" not in c.lower()
            and isinstance(gl, (int, float))
            and isinstance(gv, (int, float))
        ):
            result.matches.append(
                ParsedMatch(
                    grupo=current_group,
                    local=_clean(c),
                    visitante=_clean(f),
                    goles_local=int(gl),
                    goles_visitante=int(gv),
                )
            )

        # ¿Pronóstico de puesto? ("1 er puesto", etc.)
        b = vals.get("B")
        if isinstance(b, str) and "puesto" in b.lower():
            m = re.search(r"(\d+)", b)
            equipo = _clean(vals.get("F"))
            puntos = vals.get("E")
            if m and equipo:
                result.posiciones.append(
                    ParsedPosition(
                        posicion=int(m.group(1)),
                        equipo=equipo,
                        puntos=int(puntos) if isinstance(puntos, (int, float)) else 0,
                    )
                )

    return result
